from django.shortcuts import render, redirect
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
import django_tables2 as tables
from django_tables2.utils import A
import openpyxl


from .models import Product, File, Category
from .forms import ProductForm, FileUploadForm
from .scrapers import rtveuro_scraper, mediamarkt_scraper, mediaexpert_scraper

def index(request):
    return render(request, 'scraper/index.html')



class ProductsListTable(tables.Table):
    plu_num = tables.Column()
    art_name = tables.Column()
    category = tables.Column(accessor='category.category_name')
    auchan_price = tables.Column()
    rtveuro_url_valid = tables.BooleanColumn(verbose_name='RTV')
    rtveuro_price = tables.Column()
    mediamarkt_url_valid = tables.BooleanColumn(verbose_name='MM')
    mediamarkt_price = tables.Column()
    mediaexpert_url_valid = tables.BooleanColumn(verbose_name='ME')
    mediaexpert_price = tables.Column()
    edit_link = tables.LinkColumn('scraper:Edycja produktu', args=[A('pk')], text='Edytuj', empty_values=(), verbose_name='Edycja')

    class Meta:

        attrs = {'class': 'blueTable', 'id': 'productTable'}
        row_attrs = {'id': lambda record: record.pk}

@login_required(login_url='/login')
def products_list(request):

    products = Product.objects.all()

    table = ProductsListTable(products)
    tables.RequestConfig(request).configure(table)

    return render(request, 'scraper/products_list.html', {'products_list': table})

@login_required(login_url='/login')
def add_product(request):

    if request.method == 'POST':

        form = ProductForm(request.POST)

        if form.is_valid():
            product = form.save()
            product.save()
            return redirect('scraper:Dodaj produkt')

    else:
        form =  ProductForm()

    return render(request, 'scraper/add_product.html', {'form': form})

@login_required(login_url='/login')
def product_view(request, pk):

    product = get_object_or_404(Product, pk=pk)

    if request.method == 'POST':

        form = ProductForm(request.POST, instance=product)

        if form.is_valid():
            product = form.save()
            product.save()
            return redirect('scraper:Lista produktów')

    else:

        form = ProductForm(instance=product)

    return render(request, 'scraper/product_view.html', {'form': form})

@login_required(login_url='/login')
def scrap_prices(request):

    import time
    start_time = time.clock()

    products = Product.objects.all()

    for prod in products:
        if prod.rtveuro_url:
            prod.rtveuro_price = rtveuro_scraper(prod.rtveuro_url)
        if prod.mediamarkt_url:
            prod.mediamarkt_price = mediamarkt_scraper(prod.mediamarkt_url)
        if prod.mediaexpert_url:
            prod.mediaexpert_price = mediaexpert_scraper(prod.mediaexpert_url)
        prod.save()

    end_time = time.clock()

    time = end_time - start_time
    time = str(time).split('.', 1)[0]

    return render(request, 'scraper/scrap_prices.html', {'time': time})


def upload_auchan_prices(request):

    """
    Ceny dyskont muszą być skopiowane i wklejone do nowego pliku .xlsx
    Kolumny:
        T = CSB
        U = CSP
        Z = Nr art.
        AA = Nazwa art.
    """
    if request.method == 'POST':

        fileuploadform = FileUploadForm(request.POST, request.FILES)

        if fileuploadform.is_valid():
            fileuploadform.save()

        prices = fileuploadform.instance


        wb = openpyxl.load_workbook(prices.file.path)

        ws = wb[wb.sheetnames[0]]

        num = ''
        price = ''
        name = ''

        for row in range(ws.max_row):
            for column in "U":
                cell_name = "{}{}".format(column, row + 1)
                price = str(ws[cell_name].value)
                if ws[cell_name].value is None:
                    for new_column in 'T':
                        cell_name = "{}{}".format(new_column, row + 1)
                        price = str(ws[cell_name].value)
            for column in "Z":
                cell_name = "{}{}".format(column, row + 1)
                num = str(ws[cell_name].value)
            for column in ["AA"]:
                cell_name = "{}{}".format(column, row + 1)
                name = str(ws[cell_name].value)


            try:
                prod = Product.objects.get(plu_num=num, auchan_name=name.strip())
                prod.auchan_price = float(price)
                prod.save()
            except Product.DoesNotExist:
                continue


        return render(request, 'scraper/upload_auchan_prices_done.html')


    else:

        file = FileUploadForm()

    return render(request, 'scraper/upload_auchan_prices.html', {'file': file})

